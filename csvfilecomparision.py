import openpyxl
class filecomparison:
    def comare(self):
        file1 = '3EAH-02 DRR try (1).xlsx'
        file2 = '3EAH-02 IWB try (1).xlsx'
        wb = openpyxl.load_workbook(file1)
        wb1 = openpyxl.load_workbook(file2)
        ws = wb["vamsi"]
        ws1 = wb1["vamsi1"]
        ws2 = wb.create_sheet("difference")
        row_count = ws.max_row
        row_count1 = ws1.max_row
        list11 = []
        list1 = []
        for i in range(2,row_count):
            list11.append(ws.cell(row=i,column = 1).value)
        for j in range(2,row_count1):
            list1.append(ws1.cell(row=i,column = 2).value)

        difference = list(set(list11)-set(list1))

        print(difference)

        for i in range(0,len(difference)):
            ws2.cell(row=i+1, column=2).value = difference[i]

        wb.save('3EAH-02 DRR try (1).xlsx')





if __name__ == '__main__':
    fi = filecomparison()
    fi.comare()
